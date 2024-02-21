from openpyxl import load_workbook
import os

wb = load_workbook("./original/original_sheet.xlsx")
sheet_to_focus = "a"
for s in range(len(wb.sheetnames)):
    if wb.sheetnames[s] == sheet_to_focus:
        break
wb.active = s

wb.save("./original/original_sheet.xlsx")

os.chdir("./original")
os.system("start excel.exe original_sheet.xlsx")
